import logging
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess
import sys
from pathlib import Path
from typing import List

from config import DOSSIER_BRUT, DOSSIER_SORTIE
from core.detecteur import est_banque_internet, est_alpilink, est_compta_internet
from ui.progression import ProgressionWindow

logger = logging.getLogger("ui.justification_view")


class JustificationView:
    def __init__(self, root: tk.Tk, back):
        self.root = root
        self.back = back
        self.fichiers_detectes = {
            "banque":   [],
            "alpilink": [],
            "compta":   None
        }
        self._progression_win = None
        self.root.title("Justification Compte Internet - DLM")
        self.root.geometry("1000x800")
        self._creer_interface()
        self.rafraichir_liste()

    # ─────────────────────────────────────────────────────────────────────
    def _creer_interface(self):
        self.root.configure(bg="#f5f5f5")
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Title.TLabel",
                        font=("Segoe UI", 14, "bold"), foreground="#2c3e50")
        style.configure("TFrame",      background="#f5f5f5")
        style.configure("TLabelframe", background="#f5f5f5", foreground="#2c3e50")
        style.configure("TButton",     padding=5, font=("Segoe UI", 9))
        style.configure("TLabel",      background="#f5f5f5", foreground="#2c3e50")
        style.map("TButton", background=[("active", "#3498db")])

        for w in self.root.winfo_children():
            w.destroy()

        # ── HEADER ──────────────────────────────────────────────────────
        frame_header = ttk.Frame(self.root, padding="20 15 20 15")
        frame_header.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(frame_header, text="← Accueil",
                   command=self.back).pack(side=tk.LEFT)
        ttk.Label(frame_header,
                  text="🔍 Justification Compte Internet - DLM",
                  style="Title.TLabel").pack(side=tk.LEFT, padx=15)

        # ── CHEMINS ─────────────────────────────────────────────────────
        frame_chemins = ttk.LabelFrame(self.root,
                                       text="📁 Chemins de travail",
                                       padding="15 10")
        frame_chemins.pack(fill=tk.X, padx=15, pady=5)

        ttk.Label(frame_chemins,
                  text=f"Dossier brut   :  {DOSSIER_BRUT}",
                  foreground="#2980b9").pack(anchor=tk.W)
        ttk.Label(frame_chemins,
                  text=f"Dossier sortie :  {DOSSIER_SORTIE}",
                  foreground="#2980b9").pack(anchor=tk.W)

        # ── FICHIERS DÉTECTÉS ────────────────────────────────────────────
        frame_fichiers = ttk.LabelFrame(self.root,
                                        text="📄 Fichiers détectés",
                                        padding="15 10")
        frame_fichiers.pack(fill=tk.X, padx=15, pady=5)

        def _ligne(parent, libelle):
            f = ttk.Frame(parent)
            f.pack(fill=tk.X, pady=2)
            ttk.Label(f, text=libelle, width=18).pack(side=tk.LEFT)
            lbl = ttk.Label(f, text="—", foreground="#7f8c8d")
            lbl.pack(side=tk.LEFT)
            return lbl

        self.label_banque = _ligne(frame_fichiers, "Banque :")
        self.label_alpi   = _ligne(frame_fichiers, "Alpilink :")
        self.label_compta = _ligne(frame_fichiers, "Compta :")

        ttk.Button(frame_fichiers, text="🔄 Rafraîchir",
                   command=self.rafraichir_liste).pack(anchor=tk.E, pady=(5, 0))

        # ── JOURNAL ─────────────────────────────────────────────────────
        frame_log = ttk.LabelFrame(self.root,
                                   text="📋 Journal",
                                   padding="10 5")
        frame_log.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        self.text_log = tk.Text(frame_log, height=8, state=tk.DISABLED,
                                font=("Consolas", 9), bg="#1e1e1e", fg="#dcdcdc",
                                insertbackground="white")
        scroll_log = ttk.Scrollbar(frame_log, command=self.text_log.yview)
        self.text_log.configure(yscrollcommand=scroll_log.set)
        scroll_log.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_log.pack(fill=tk.BOTH, expand=True)

        # ── TABLEAU RÉSULTATS ────────────────────────────────────────────
        frame_tableau = ttk.LabelFrame(self.root,
                                       text="📊 Résultats",
                                       padding="10 5")
        frame_tableau.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        colonnes = (
            "commande", "montant_compta", "montant_banque",
            "montant_alpi", "ecart_banque", "ecart_alpi",
            "statut", "date", "source"
        )
        self.tableau = ttk.Treeview(frame_tableau,
                                    columns=colonnes,
                                    show="headings",
                                    height=10)
        for col in colonnes:
            self.tableau.heading(col, text=col.replace("_", " ").title())
            self.tableau.column(col, width=120, anchor=tk.CENTER)

        scroll_tab = ttk.Scrollbar(frame_tableau,
                                   orient=tk.VERTICAL,
                                   command=self.tableau.yview)
        self.tableau.configure(yscrollcommand=scroll_tab.set)
        scroll_tab.pack(side=tk.RIGHT, fill=tk.Y)

        scroll_h = ttk.Scrollbar(frame_tableau,
                                  orient=tk.HORIZONTAL,
                                  command=self.tableau.xview)
        self.tableau.configure(xscrollcommand=scroll_h.set)
        scroll_h.pack(side=tk.BOTTOM, fill=tk.X)
        self.tableau.pack(fill=tk.BOTH, expand=True)

        self.tableau.tag_configure("non_validee",     background="#FADBD8")
        self.tableau.tag_configure("libelle_inconnu", background="#FEF9E7")

        # ── STATUT ET BOUTONS ────────────────────────────────────────────
        frame_statut = ttk.Frame(self.root, padding="15 10")
        frame_statut.pack(fill=tk.X)

        self.label_statut = tk.Label(
            frame_statut,
            text="⏳ Prêt à démarrer",
            foreground="#7f8c8d",
            font=("Segoe UI", 9, "italic"),
            bg="#f5f5f5"
        )
        self.label_statut.pack(side=tk.LEFT)

        frame_btn = ttk.Frame(frame_statut)
        frame_btn.pack(side=tk.RIGHT)

        self.btn_lancer = ttk.Button(frame_btn,
                                     text="▶ Lancer la justification",
                                     command=self.lancer_justification)
        self.btn_lancer.pack(side=tk.LEFT, padx=5)

        ttk.Button(frame_btn, text="📥 Exporter Excel",
                   command=self.exporter_excel).pack(side=tk.LEFT, padx=5)

        ttk.Button(frame_btn, text="📂 Dossier sortie",
                   command=lambda: self._ouvrir_dossier(DOSSIER_SORTIE)
                   ).pack(side=tk.LEFT, padx=5)

    # ─────────────────────────────────────────────────────────────────────
    def _ouvrir_dossier(self, dossier: Path):
        try:
            if sys.platform == "win32":
                subprocess.Popen(["explorer", str(dossier)])
            else:
                subprocess.Popen(["xdg-open", str(dossier)])
        except Exception as e:
            self.ajouter_log(f"Impossible d'ouvrir le dossier: {e}", "ERROR")

    # ─────────────────────────────────────────────────────────────────────
    def ajouter_log(self, message: str, niveau: str = "INFO"):
        niveau = niveau.upper()
        prefix = {"SUCCESS": "✅", "ERROR": "❌",
                  "WARNING": "⚠️", "INFO": "ℹ️"}.get(niveau, "ℹ️")
        log_fn = getattr(logger, niveau.lower(), logger.info)
        log_fn(f"{prefix} {message}")
        self.root.after(0, lambda: self._mettre_a_jour_log_ui(prefix, message, niveau))

    def _mettre_a_jour_log_ui(self, prefix: str, message: str, niveau: str):
        if niveau in ("ERROR", "SUCCESS"):
            self.label_statut.config(
                text=f"{prefix} {message}",
                foreground="#e74c3c" if niveau == "ERROR" else "#27ae60"
            )
        self.text_log.configure(state=tk.NORMAL)
        self.text_log.insert(tk.END, f"{prefix} {message}\n")
        self.text_log.see(tk.END)
        self.text_log.configure(state=tk.DISABLED)

    # ─────────────────────────────────────────────────────────────────────
    def rafraichir_liste(self):
        try:
            self.btn_lancer.config(state=tk.DISABLED)
            self.label_statut.config(text="🔍 Détection des fichiers en cours…",
                                     foreground="#f39c12")
            self.root.update()

            self.fichiers_detectes = {"banque": [], "alpilink": [], "compta": None}
            trouves = 0

            for f in DOSSIER_BRUT.glob("*.*"):
                try:
                    if est_banque_internet(f):
                        self.fichiers_detectes["banque"].append(f)
                        trouves += 1
                    elif est_alpilink(f):
                        self.fichiers_detectes["alpilink"].append(f)
                        trouves += 1
                    elif est_compta_internet(f):
                        self.fichiers_detectes["compta"] = f
                        trouves += 1
                except Exception:
                    pass

            self.label_banque.config(
                text=", ".join(f.name for f in self.fichiers_detectes["banque"]) or "—"
            )
            self.label_alpi.config(
                text=", ".join(f.name for f in self.fichiers_detectes["alpilink"]) or "—"
            )
            self.label_compta.config(
                text=self.fichiers_detectes["compta"].name
                if self.fichiers_detectes["compta"] else "—"
            )

            if trouves > 0:
                self.btn_lancer.config(state=tk.NORMAL)
                self.ajouter_log(
                    f"Détection terminée : {trouves} fichier(s) trouvé(s)", "SUCCESS"
                )
            else:
                self.label_statut.config(text="⚠️ Aucun fichier détecté",
                                         foreground="#e67e22")

        except Exception as e:
            self.ajouter_log(f"Erreur lors de la détection : {e}", "ERROR")
            self.btn_lancer.config(state=tk.NORMAL)

    # ─────────────────────────────────────────────────────────────────────
    def lancer_justification(self):
        self.btn_lancer.config(state=tk.DISABLED)
        self.label_statut.config(text="⏳ Justification en cours…",
                                 foreground="#f39c12")

        # Vider le tableau
        for row in self.tableau.get_children():
            self.tableau.delete(row)

        try:
            from core.justification_handler import JustificationHandler

            fichiers = (
                self.fichiers_detectes["banque"]
                + self.fichiers_detectes["alpilink"]
                + ([self.fichiers_detectes["compta"]]
                   if self.fichiers_detectes["compta"] else [])
            )

            # ── Ouvrir la barre de progression ──────────────────────────
            # total = nb fichiers + 4 étapes internes
            self._progression_win = ProgressionWindow(self.root, len(fichiers) + 4)

            handler = JustificationHandler(
                callback_log=lambda msg: self.root.after(
                    0, lambda m=msg: self.ajouter_log(m, "INFO")
                ),
                callback_fin=lambda *args: self.root.after(
                    0, lambda a=args: self._on_justification_terminee(*a)
                ),
                # ── nouveau callback progression ─────────────────────────
                callback_progression=lambda val, txt: self.root.after(
                    0, lambda v=val, t=txt: self._maj_progression(v, t)
                )
            )
            handler.lancer(fichiers, str(DOSSIER_SORTIE))

        except Exception as e:
            # Fermer la barre si elle est ouverte
            if self._progression_win:
                self._progression_win.fermer()
                self._progression_win = None
            self.ajouter_log(f"Erreur pendant la justification: {e}", "ERROR")
            self.btn_lancer.config(state=tk.NORMAL)

    # ─────────────────────────────────────────────────────────────────────
    def _maj_progression(self, valeur: int, texte: str):
        """Met à jour la fenêtre de progression (appelé via root.after)."""
        if self._progression_win:
            try:
                self._progression_win.maj(valeur, texte)
            except Exception:
                pass  # fenêtre déjà fermée

    # ─────────────────────────────────────────────────────────────────────
    def _on_justification_terminee(
            self,
            ecarts: list,
            non_just: list,
            erreurs_fmt: list,
            toutes_validees: list,
            tous_non_valides: list
    ):
        # ── Fermer la barre de progression ──────────────────────────────
        if self._progression_win:
            self._progression_win.fermer()
            self._progression_win = None

        self.btn_lancer.config(state=tk.NORMAL)

        # ── Non justifiées ───────────────────────────────────────────────
        for nj in non_just:
            self.tableau.insert("", tk.END, values=(
                nj.get("commande"),
                nj.get("montant_compta"),
                nj.get("montant_banque"),
                nj.get("montant_alpi"),
                nj.get("ecart_banque"),
                nj.get("ecart_alpi"),
                nj.get("statut"),
                nj.get("date"),
                nj.get("source"),
            ), tags=("non_validee",))

        # ── Libellés inconnus ────────────────────────────────────────────
        inconnus = [e for e in erreurs_fmt if "non reconnu" in e.get("Erreur", "")]
        for inc in inconnus:
            self.tableau.insert("", tk.END, values=(
                inc.get("Libellé", ""),
                inc.get("Montant", ""),
                "", "", "", "",
                "libellé inconnu",
                inc.get("Date", ""),
                inc.get("Fichier", ""),
            ), tags=("libelle_inconnu",))

        self.ajouter_log(
            f"Terminé — {len(toutes_validees)} justifiées | "
            f"{len(non_just)} non justifiées | "
            f"{len(inconnus)} libellés inconnus",
            "SUCCESS"
        )

    # ─────────────────────────────────────────────────────────────────────
    def exporter_excel(self):
        lignes = self.tableau.get_children()
        if not lignes:
            messagebox.showwarning("Attention", "Aucune donnée à exporter.")
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            messagebox.showerror("Erreur",
                                 "openpyxl non installé.\nExécutez: pip install openpyxl")
            return

        chemin = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=str(DOSSIER_SORTIE),
            initialfile="justification_internet.xlsx"
        )
        if not chemin:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Justification"

            entetes = [
                "Commande", "Montant Compta", "Montant Banque",
                "Montant Alpi", "Écart Banque", "Écart Alpi",
                "Statut", "Date", "Source"
            ]
            ws.append(entetes)
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="2980B9")
                cell.alignment = Alignment(horizontal="center")

            couleurs = {
                "non justifiée":   "FADBD8",
                "libellé inconnu": "FEF9E7",
            }

            for row_id in lignes:
                vals   = self.tableau.item(row_id)["values"]
                statut = vals[6] if len(vals) > 6 else ""
                ws.append(list(vals))
                fill_color = couleurs.get(statut, "FFFFFF")
                for cell in ws[ws.max_row]:
                    cell.fill      = PatternFill("solid", fgColor=fill_color)
                    cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_w = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = max_w + 4

            wb.save(chemin)
            self.ajouter_log(f"Export Excel : {chemin}", "SUCCESS")
            messagebox.showinfo("Export réussi",
                                f"Fichier exporté avec succès :\n{chemin}")

        except Exception as e:
            self.ajouter_log(f"Erreur export Excel: {e}", "ERROR")
            messagebox.showerror("Erreur", f"Échec de l'export : {e}")
