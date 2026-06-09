"""
Module 2 - Handler de justification
Rapproche COMPTA / BANQUE / ALPILINK
"""

import threading
import pandas as pd
from openpyxl import Workbook
from pathlib import Path
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import re

from config import logger
from core.utils.colonnes import RE_COMMANDE

from modules.module_2.handlers.compta_handler import (
    charger_compta,
    extraire_commandes as extraire_commandes_compta,
)
from modules.module_2.handlers.banque_handler import (
    charger_banque,
    extraire_commandes as extraire_commandes_banque,
)
from modules.module_2.handlers.alpilink_handler import (
    charger_alpilink,
    extraire_commandes_alpilink,
)

# ==========================================================
# CONSTANTES
# ==========================================================

RE_COMMANDE_PATTERN = (
    re.compile(RE_COMMANDE) if isinstance(RE_COMMANDE, str) else RE_COMMANDE
)

# Couleurs Excel
FILL_VERT   = PatternFill("solid", fgColor="C6EFCE")
FILL_JAUNE  = PatternFill("solid", fgColor="FFEB9C")
FILL_ROUGE  = PatternFill("solid", fgColor="FFC7CE")
FILL_ORANGE = PatternFill("solid", fgColor="FFD966")
FONT_GRAS   = Font(bold=True)

# ==========================================================
# CLASSE PRINCIPALE
# ==========================================================

class JustificationHandler:
    def __init__(
        self,
        callback_log=None,
        callback_fin=None,
        callback_progression=None,
    ):
        self.lock = threading.Lock()
        self.log_messages = []
        self.callback_log = callback_log
        self.callback_fin = callback_fin
        self.callback_progression = callback_progression

    # ------------------------------------------------------------------
    def log(self, msg: str):
        with self.lock:
            self.log_messages.append(msg)
            logger.info(f"[JUSTIF] {msg}")
            if self.callback_log:
                self.callback_log(msg)

    def _progression(self, current: int, msg: str):
        if self.callback_progression:
            self.callback_progression(current, msg)

    # ------------------------------------------------------------------
    @staticmethod
    def _detecter_type(chemin: Path) -> str:
        try:
            ext = chemin.suffix.lower()
            if ext == ".xlsx":
                df = pd.read_excel(chemin, engine="openpyxl", nrows=5, dtype=str)
            elif ext == ".xls":
                df = pd.read_excel(chemin, engine="xlrd", nrows=5, dtype=str)
            elif ext == ".csv":
                df = pd.read_csv(
                    chemin, sep=None, engine="python",
                    nrows=5, dtype=str, encoding="utf-8"
                )
            else:
                return "inconnu"

            if df is None or df.empty:
                return "inconnu"

            cols = set(df.columns.str.strip().str.lower())

            compta_cols = {"montant signé", "journal", "sens", "lettrage"}
            if compta_cols.issubset(cols):
                return "compta"

            banque_cols = {"date du paiement", "commande", "montant du paiement"}
            if banque_cols.issubset(cols):
                return "banque"

            alpi_cols = {"id commande", "prix total", "canal de vente"}
            if alpi_cols.issubset(cols):
                return "alpilink"

            return "inconnu"

        except Exception as e:
            logger.debug(f"[JUSTIF] Erreur détection {chemin.name} : {e}")
            return "inconnu"

    # ------------------------------------------------------------------
    @staticmethod
    def _rapprocher_commandes(
        map_compta: dict,
        commandes_banque: list,
        commandes_alpilink: list,
    ):
        logger.info(
            f"[JUSTIF] Rapprochement : {len(map_compta)} compta | "
            f"{len(commandes_banque)} banque | "
            f"{len(commandes_alpilink)} alpilink"
        )

        map_banque   = {c["commande"]: c for c in commandes_banque}   if commandes_banque   else {}
        map_alpilink = {c["commande"]: c for c in commandes_alpilink} if commandes_alpilink else {}

        logger.info(f"[JUSTIF][DEBUG] Clés compta   (10) : {list(map_compta.keys())[:10]}")
        logger.info(f"[JUSTIF][DEBUG] Clés banque   (10) : {list(map_banque.keys())[:10]}")
        logger.info(f"[JUSTIF][DEBUG] Clés alpilink (10) : {list(map_alpilink.keys())[:10]}")

        validees    = []
        non_valides = []

        for num_cmd, cmd_compta in map_compta.items():
            montant_compta = float(cmd_compta.get("montant", 0) or 0)
            cmd_banque     = map_banque.get(num_cmd)
            cmd_alpilink   = map_alpilink.get(num_cmd)

            trouve_banque   = cmd_banque   is not None
            trouve_alpilink = cmd_alpilink is not None

            record = {
                "commande":       num_cmd,
                "montant_compta": montant_compta,
                "date":           cmd_compta.get("date", ""),
                "journal":        cmd_compta.get("journal", ""),
            }

            if trouve_banque and trouve_alpilink:
                montant_banque   = float(cmd_banque.get("montant", 0)   or 0)
                montant_alpilink = float(cmd_alpilink.get("montant", 0) or 0)
                record.update({
                    "montant_banque":   montant_banque,
                    "montant_alpilink": montant_alpilink,
                    "ecart_banque":     round(abs(montant_compta - montant_banque),   2),
                    "ecart_alpilink":   round(abs(montant_compta - montant_alpilink), 2),
                    "source":           "banque + alpilink",
                    "statut":           "justifiée (banque + alpilink)",
                    "email":            cmd_banque.get("email", ""),
                    "type_paiement":    cmd_banque.get("type_paiement", ""),
                    "contrat":          cmd_banque.get("contrat", ""),
                    "n_remise":         cmd_banque.get("n_remise", ""),
                    "date_remise":      cmd_banque.get("date_remise", ""),
                    "canal":            cmd_alpilink.get("canal", ""),
                    "statut_alpi":      cmd_alpilink.get("statut", ""),
                })
                validees.append(record)

            elif trouve_banque:
                montant_banque = float(cmd_banque.get("montant", 0) or 0)
                record.update({
                    "montant_banque":   montant_banque,
                    "montant_alpilink": None,
                    "ecart_banque":     round(abs(montant_compta - montant_banque), 2),
                    "ecart_alpilink":   None,
                    "source":           "banque",
                    "statut":           "justifiée (banque)",
                    "email":            cmd_banque.get("email", ""),
                    "type_paiement":    cmd_banque.get("type_paiement", ""),
                    "contrat":          cmd_banque.get("contrat", ""),
                    "n_remise":         cmd_banque.get("n_remise", ""),
                    "date_remise":      cmd_banque.get("date_remise", ""),
                    "canal":            "",
                    "statut_alpi":      "",
                })
                validees.append(record)

            elif trouve_alpilink:
                montant_alpilink = float(cmd_alpilink.get("montant", 0) or 0)
                record.update({
                    "montant_banque":   None,
                    "montant_alpilink": montant_alpilink,
                    "ecart_banque":     None,
                    "ecart_alpilink":   round(abs(montant_compta - montant_alpilink), 2),
                    "source":           "alpilink",
                    "statut":           "justifiée (alpilink)",
                    "email":            "",
                    "type_paiement":    "",
                    "contrat":          "",
                    "n_remise":         "",
                    "date_remise":      "",
                    "canal":            cmd_alpilink.get("canal", ""),
                    "statut_alpi":      cmd_alpilink.get("statut", ""),
                })
                validees.append(record)

            else:
                record.update({
                    "montant_banque":   None,
                    "montant_alpilink": None,
                    "ecart_banque":     None,
                    "ecart_alpilink":   None,
                    "source":           "compta uniquement",
                    "statut":           "non justifiée",
                    "email":            "",
                    "type_paiement":    "",
                    "contrat":          "",
                    "n_remise":         "",
                    "date_remise":      "",
                    "canal":            "",
                    "statut_alpi":      "",
                })
                non_valides.append(record)

        return validees, non_valides

    # ------------------------------------------------------------------
    @staticmethod
    def _ecrire_onglet(ws, lignes: list, fill_defaut: PatternFill, nom: str):
        if not lignes:
            ws.append(["(aucune ligne)"])
            return

        entetes = list(lignes[0].keys())
        ws.append(entetes)
        for cell in ws[1]:
            cell.font = FONT_GRAS

        for i, ligne in enumerate(lignes):
            try:
                ws.append(list(ligne.values()))
                row_ws = ws[ws.max_row]

                if fill_defaut == FILL_VERT:
                    ecart_b = ligne.get("ecart_banque")   or 0
                    ecart_a = ligne.get("ecart_alpilink") or 0
                    fill = FILL_JAUNE if (ecart_b > 0 or ecart_a > 0) else FILL_VERT
                else:
                    fill = fill_defaut

                for cell in row_ws:
                    cell.fill = fill

            except Exception as e:
                logger.error(f"[JUSTIF][EXPORT] Erreur ligne {i} onglet {nom} : {e}")

    # ------------------------------------------------------------------
    @staticmethod
    def _exporter_excel(validees, non_valides, erreurs_fmt, dossier_sortie):
        dossier = Path(dossier_sortie)
        dossier.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        chemin    = dossier / f"justification_{timestamp}.xlsx"

        try:
            wb = Workbook()

            ws1 = wb.active
            ws1.title = "Validées"
            JustificationHandler._ecrire_onglet(ws1, validees, FILL_VERT, "Validées")

            ws2 = wb.create_sheet("Non justifiées")
            JustificationHandler._ecrire_onglet(ws2, non_valides, FILL_ROUGE, "Non justifiées")

            erreurs_libelles = [
                e for e in erreurs_fmt
                if "non reconnu" in e.get("Erreur", "")
            ]
            erreurs_autres = [
                e for e in erreurs_fmt
                if "non reconnu" not in e.get("Erreur", "")
            ]

            ws3 = wb.create_sheet("Libellés inconnus")
            JustificationHandler._ecrire_onglet(ws3, erreurs_libelles, FILL_ORANGE, "Libellés inconnus")

            ws4 = wb.create_sheet("Erreurs format")
            JustificationHandler._ecrire_onglet(ws4, erreurs_autres, FILL_ORANGE, "Erreurs format")

            wb.save(chemin)
            logger.info(f"[JUSTIF][EXPORT] ✓ Fichier généré : {chemin}")
            return str(chemin)

        except Exception as e:
            logger.error(f"[JUSTIF][EXPORT] ❌ Erreur export : {e}", exc_info=True)
            return None

    # ------------------------------------------------------------------
    def _run(self, fichiers, dossier_sortie):
        self.log("═" * 60)
        self.log("DEBUT JUSTIFICATION")
        self.log("═" * 60)

        chemin_compta   = None
        chemins_banques = []
        chemins_alpis   = []
        erreurs_fmt     = []

        # ── PHASE 1 : CHARGEMENT ──────────────────────────────────────
        total = len(fichiers)
        self.log(f"\n[PHASE 1] CHARGEMENT ({total} fichiers)")
        self.log("─" * 60)

        for i, chemin in enumerate(fichiers, 1):
            chemin_path = Path(chemin)
            nom = chemin_path.name
            self._progression(i, f"Chargement : {nom}")

            try:
                type_ = self._detecter_type(chemin_path)
                self.log(f"  ✅ [{i}/{total}] {nom} — type : {type_.upper()}")

                if type_ == "compta":
                    chemin_compta = chemin_path
                elif type_ == "banque":
                    chemins_banques.append(chemin_path)
                elif type_ == "alpilink":
                    chemins_alpis.append(chemin_path)
                else:
                    erreurs_fmt.append({"Fichier": nom, "Erreur": "Type de fichier inconnu"})
                    self.log(f"  ❌ [{i}/{total}] {nom} — type inconnu")

            except Exception as e:
                erreurs_fmt.append({"Fichier": nom, "Erreur": str(e)})
                logger.error(f"[JUSTIF] Erreur détection {nom} : {e}", exc_info=True)

        if chemin_compta is None:
            self.log("\n❌ ERREUR : Aucun fichier COMPTA détecté — abandon")
            if self.callback_fin:
                self.callback_fin([], [], erreurs_fmt, [], [])
            return

        # ── PHASE 2 : EXTRACTION ──────────────────────────────────────
        self.log(f"\n[PHASE 2] EXTRACTION")
        self.log("─" * 60)
        self._progression(total + 1, "Extraction…")

        map_compta         = {}
        commandes_banque   = []
        commandes_alpilink = []

        # COMPTA
        try:
            df_compta = charger_compta(chemin_compta)
            map_compta, errs_compta = extraire_commandes_compta(df_compta)
            erreurs_fmt.extend(errs_compta)
            self.log(f"  COMPTA   : {len(map_compta)} commandes | {len(errs_compta)} libellés inconnus")
        except Exception as e:
            logger.error(f"[JUSTIF] Erreur extraction compta : {e}", exc_info=True)
            self.log(f"  ❌ COMPTA : Erreur extraction — {e}")

        # BANQUES
        for j, chemin_banque in enumerate(chemins_banques, 1):
            try:
                df_banque = charger_banque(chemin_banque)
                result    = extraire_commandes_banque(df_banque)
                if isinstance(result, tuple):
                    commandes_banque.extend(result[0])
                    erreurs_fmt.extend(result[1])
                else:
                    commandes_banque.extend(result)
            except Exception as e:
                logger.error(f"[JUSTIF] Erreur extraction banque {j} : {e}", exc_info=True)
                self.log(f"  ❌ BANQUE {j} : Erreur extraction — {e}")

        self.log(f"  BANQUES  : {len(commandes_banque)} commandes")

        # ALPILINK
        for j, chemin_alpi in enumerate(chemins_alpis, 1):
            try:
                df_normal, df_buyclub = charger_alpilink(chemin_alpi)
                result = extraire_commandes_alpilink(df_normal, df_buyclub)
                if isinstance(result, tuple):
                    commandes_alpilink.extend(result[0])
                    erreurs_fmt.extend(result[1])
                else:
                    commandes_alpilink.extend(result)
            except Exception as e:
                logger.error(f"[JUSTIF] Erreur extraction alpilink {j} : {e}", exc_info=True)
                self.log(f"  ❌ ALPILINK {j} : Erreur extraction — {e}")

        self.log(f"  ALPILINK : {len(commandes_alpilink)} commandes")

        # ── PHASE 3 : RAPPROCHEMENT ───────────────────────────────────
        self.log(f"\n[PHASE 3] RAPPROCHEMENT")
        self.log("─" * 60)
        self._progression(total + 2, "Rapprochement…")

        validees, non_valides = self._rapprocher_commandes(
            map_compta, commandes_banque, commandes_alpilink
        )

        self.log(f"  Validées     : {len(validees)}")
        self.log(f"  Non valides  : {len(non_valides)}")

        # ── PHASE 4 : EXPORT ──────────────────────────────────────────
        self.log(f"\n[PHASE 4] EXPORT")
        self.log("─" * 60)
        self._progression(total + 3, "Export Excel…")

        chemin_export = self._exporter_excel(validees, non_valides, erreurs_fmt, dossier_sortie)

        if chemin_export:
            self.log(f"  ✓ Rapport : {chemin_export}")
        else:
            self.log("  ❌ Échec export Excel")

        self._progression(total + 4, "Terminé ✓")

        self.log("\n" + "═" * 60)
        self.log("FIN JUSTIFICATION")
        self.log("═" * 60)

        # ── CORRECTION : 5 arguments attendus par la vue ──────────────
        if self.callback_fin:
            self.callback_fin(
                validees,      # ecarts          (arg 1)
                non_valides,   # non_just         (arg 2)
                erreurs_fmt,   # erreurs_fmt      (arg 3)
                validees,      # toutes_validees  (arg 4)
                non_valides,   # tous_non_valides (arg 5)
            )

    # ------------------------------------------------------------------
    def run(self, fichiers, dossier_sortie):
        thread = threading.Thread(
            target=self._run,
            args=(fichiers, dossier_sortie),
            daemon=True,
        )
        thread.start()
        return thread

    def lancer(self, fichiers, dossier_sortie):
        return self.run(fichiers, dossier_sortie)

# ==========================================================
# EXPORT
# ==========================================================
__all__ = ["JustificationHandler"]
