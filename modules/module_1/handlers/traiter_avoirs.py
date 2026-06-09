"""
Module 1 - Handler AVOIRS
Logique métier identique à l'original, adaptée au nouveau core.
"""

from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import load_workbook

from config import DOSSIER_SORTIE
from config import logger
from core.moniteur_schema import comparer_schema
from core.utils.montant import format_montant_compta
from core.utils.constantes import (
    STE_DLM,
    COMPTE_TRANSIT,
    COMPTE_AVOIR,
    JOURNAL_OD,
    ANALYTIQUE_VIDE,
    COL_STE, COL_DATE, COL_COMPTE, COL_AUX,
    COL_PIECE, COL_OBJET, COL_DEBIT, COL_CREDIT,
    COL_JOURNAL, COL_ANALYTIQUE,
    COLONNES_SORTIE,
)

# ==========================================================
# EXCEPTION MÉTIER
# ==========================================================

class NotAvoirFileError(Exception):
    """Levée si aucun avoir exploitable n'est détecté."""
    pass

# ==========================================================
# INDEX DES COLONNES (base 0)
# ==========================================================
_COL_NOM            = 1   # B
_COL_PRENOM         = 2   # C
_COL_DATE_CREATION  = 4   # E
_COL_DATE_EXPIR     = 5   # F
_COL_STATUT_G       = 6   # G
_COL_STATUT_H       = 7   # H  ← contient "Remboursement" ou "Avoir"
_COL_MONTANT        = 8   # I
_COL_COMMANDE       = 10  # K

# ==========================================================
# UTILITAIRES INTERNES
# ==========================================================

def _norm(val) -> str:
    """Normalise une valeur texte : strip + majuscules."""
    return str(val).strip().upper() if val is not None else ""


def _statut_to_debit_commande(statut_g, statut_h, montant) -> Optional[bool]:
    """
    Détermine si la première ligne est en DÉBIT (True) ou CRÉDIT (False).

    Logique :
    - AVOIR : toujours Crédit 580010 / Débit 580012 → False
    - REMBOURSEMENT + positif : Débit 580010 / Crédit 580012 → True
    - REMBOURSEMENT + négatif : Crédit 580010 / Débit 580012 → False
    """
    s1 = _norm(statut_g)
    s2 = _norm(statut_h)
    positif = montant >= 0

    if s2 == "AVOIR":
        return False

    if s2 == "REMBOURSEMENT":
        return positif

    # Fallback colonne G
    if s1 in ("AVOIR", "CONSUMED"):
        return False

    if s1 in ("DEDUCTED", "REMBOURSEMENT"):
        return positif

    logger.warning(f"[AVOIRS] Statut avoir inconnu : G={s1!r} H={s2!r} → ligne ignorée")
    return None


# ==========================================================
# HANDLER PRINCIPAL
# ==========================================================

def traiter_avoirs(fichier: Path) -> tuple[str, str]:
    """
    Traite un fichier d'avoirs (.xlsx) et génère les écritures comptables.

    Logique comptable :
    - Chaque avoir génère 2 écritures :
      1. COMPTE_TRANSIT   (580010DS5) : débit ou crédit selon sens
      2. COMPTE_AVOIR     (580012DS5) : crédit ou débit (sens inverse)

    Retourne:
        ("OK", chemin_fichier) | ("ERREUR", message)
    """

    fichier = Path(fichier)
    if not fichier.exists():
        msg = f"Fichier AVOIR introuvable : {fichier}"
        logger.error(f"[AVOIRS] {msg}")
        return "ERREUR", msg

    logger.info(f"[MODULE1][AVOIRS] Début traitement : {fichier.name}")

    lignes     = []
    nb_ignores = 0

    wb = load_workbook(fichier, data_only=True)

    try:
        ws = wb.active

        # ----------------------------------------------------------
        # Vérification du schéma (snapshot ligne 1 via pandas)
        # ----------------------------------------------------------
        try:
            df_schema = pd.read_excel(fichier, nrows=0, header=0)
            comparer_schema(df_schema, "avoirs")
        except ValueError as e:
            msg = f"Schéma invalide : {e}"
            logger.error(f"[AVOIRS] {msg}")
            return "ERREUR", msg

        # ----------------------------------------------------------
        # Traitement des lignes
        # ----------------------------------------------------------
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):

            # Lecture des cellules
            nom            = row[_COL_NOM].value
            prenom         = row[_COL_PRENOM].value
            date_creation  = row[_COL_DATE_CREATION].value
            date_expir     = row[_COL_DATE_EXPIR].value
            statut_g       = row[_COL_STATUT_G].value
            statut_h       = row[_COL_STATUT_H].value if len(row) > _COL_STATUT_H else None
            montant        = row[_COL_MONTANT].value
            commande       = row[_COL_COMMANDE].value

            # Ligne incomplète
            if montant is None or date_creation is None:
                logger.debug(f"[AVOIRS] Ligne {idx} ignorée : montant ou date manquant")
                nb_ignores += 1
                continue

            # Sens comptable
            debit_commande = _statut_to_debit_commande(statut_g, statut_h, montant)
            if debit_commande is None:
                nb_ignores += 1
                continue

            # Formatage
            try:
                date_str = date_creation.strftime("%d/%m/%Y")
            except AttributeError:
                logger.warning(
                    f"[AVOIRS] Ligne {idx} ignorée : date_creation non formatable ({date_creation!r})"
                )
                nb_ignores += 1
                continue

            piece       = f"JOURNEE DU {date_str}"
            montant_fmt = format_montant_compta(montant)

            exp_str = (
                date_expir.strftime("%d/%m/%Y")
                if date_expir is not None else ""
            )
            libelle_avoir = f"{_norm(nom)} {_norm(prenom)} {exp_str}".strip()
            libelle_cmd   = str(commande).strip() if commande is not None else ""

            # ✅ Ligne COMMANDE (580010DS5 - TRANSIT)
            lignes.append({
                COL_STE:        STE_DLM,
                COL_DATE:       date_str,
                COL_COMPTE:     COMPTE_TRANSIT,
                COL_AUX:        ANALYTIQUE_VIDE,
                COL_PIECE:      piece,
                COL_OBJET:      libelle_cmd,
                COL_DEBIT:      montant_fmt if debit_commande else "",
                COL_CREDIT:     "" if debit_commande else montant_fmt,
                COL_JOURNAL:    JOURNAL_OD,
                COL_ANALYTIQUE: ANALYTIQUE_VIDE,
            })

            # ✅ Ligne AVOIR (580012DS5 - AVOIR)
            lignes.append({
                COL_STE:        STE_DLM,
                COL_DATE:       date_str,
                COL_COMPTE:     COMPTE_AVOIR,
                COL_AUX:        ANALYTIQUE_VIDE,
                COL_PIECE:      piece,
                COL_OBJET:      libelle_avoir,
                COL_DEBIT:      "" if debit_commande else montant_fmt,
                COL_CREDIT:     montant_fmt if debit_commande else "",
                COL_JOURNAL:    JOURNAL_OD,
                COL_ANALYTIQUE: ANALYTIQUE_VIDE,
            })

        # ----------------------------------------------------------
        # Aucun mouvement détecté
        # ----------------------------------------------------------
        if not lignes:
            msg = f"Aucune ligne AVOIR exploitable dans {fichier.name}"
            logger.warning(f"[AVOIRS] {msg}")
            return "ERREUR", msg

        logger.info(
            f"[AVOIRS] {len(lignes) // 2} avoirs traités "
            f"({nb_ignores} lignes ignorées)"
        )

        # ----------------------------------------------------------
        # Export CSV
        # ----------------------------------------------------------
        try:
            DOSSIER_SORTIE.mkdir(parents=True, exist_ok=True)

            df_final = pd.DataFrame(lignes, columns=COLONNES_SORTIE)
            sortie   = DOSSIER_SORTIE / f"{fichier.stem}_avoirs.csv"
            df_final.to_csv(sortie, sep=";", index=False, encoding="latin-1")

            logger.info(
                f"[MODULE1][AVOIRS] Export réussi : {sortie.name} ({len(lignes)} écritures)"
            )
            return "OK", str(sortie)

        except Exception as e:
            msg = f"Erreur à l'export : {e}"
            logger.error(f"[AVOIRS] {msg}")
            return "ERREUR", msg

    except Exception as e:
        msg = f"Erreur de traitement : {e}"
        logger.error(f"[AVOIRS] {msg}")
        return "ERREUR", msg

    finally:
        wb.close()


# ==========================================================
__all__ = ["traiter_avoirs", "NotAvoirFileError"]
