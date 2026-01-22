import csv
from pathlib import Path
from openpyxl import load_workbook
from config import DOSSIER_SORTIE


# Exception levée si aucun avoir exploitable n’est détecté
class NotAvoirFileError(Exception):
    pass


def format_montant(valeur) -> str:
    """
    Formate un montant pour l’export comptable :
    - Valeur absolue
    - Deux décimales
    - Virgule comme séparateur décimal
    """
    if valeur is None:
        return ""

    if isinstance(valeur, str):
        valeur = valeur.replace(",", ".").strip()
        valeur = float(valeur)

    return f"{abs(valeur):.2f}".replace(".", ",")


def _norm(val) -> str:
    """
    Normalise une valeur texte :
    - Conversion en chaîne
    - Suppression des espaces
    - Passage en majuscules
    """
    return str(val).strip().upper() if val is not None else ""


def _statut_to_debit_commande(statut_g: str, statut_h: str) -> bool | None:
    """
    Détermine le sens comptable à partir du statut de l’avoir.

    Retourne :
      - True  → compte commande (580010DS5) en Débit,
                compte avoir (580012DS5) en Crédit
      - False → inversion des sens
      - None  → statut inconnu (ligne ignorée)
    """
    s1 = _norm(statut_g)
    s2 = _norm(statut_h)

    # Cas export français
    if s1 == "AVOIR":
        return True

    # Cas export anglais
    if s1 == "CONSUMED":
        return True
    if s1 == "DEDUCTED":
        return False

    # Cas fallback : "Remboursement" présent sur la colonne voisine
    if s1 == "REMBOURSEMENT" or s2 == "REMBOURSEMENT":
        return False

    return None


def traiter_avoirs(fichier: Path):
    """
    Traite un fichier d’avoirs et génère les écritures comptables.

    MATRICE DE SORTIE :
    STE | DATE | COMPTE | Auxiliaire | n° pièce | OBJET | D | C | Journal | Analytique
    """

    # Constantes comptables
    STE = "DLM"
    COMPTE_COMMANDE = "580010DS5"
    COMPTE_AVOIR = "580012DS5"
    AUXILIAIRE = ""
    ANALYTIQUE = ""
    JOURNAL = "CEBOOBA"

    lignes = []

    # Ouverture du fichier Excel (valeurs calculées uniquement)
    wb = load_workbook(fichier, data_only=True)
    ws = wb.active

    try:
        # Parcours des lignes (en ignorant l’en-tête)
        for row in ws.iter_rows(min_row=2):

            # Colonnes du fichier avoir
            nom = row[1].value              # B
            prenom = row[2].value           # C
            date_creation = row[4].value    # E
            date_expiration = row[5].value  # F

            statut_g = row[6].value         # G
            statut_h = row[7].value if len(row) > 7 else None  # H (fallback)

            montant = row[8].value          # I
            commande = row[10].value        # K

            # Ligne incomplète → ignorée
            if montant is None or date_creation is None:
                continue

            # Détermination du sens comptable
            debit_commande = _statut_to_debit_commande(statut_g, statut_h)
            if debit_commande is None:
                continue

            date_ecriture = date_creation.strftime("%d/%m/%Y")
            piece = f"JOURNEE DU {date_ecriture}"
            montant_fmt = format_montant(montant)

            # ------------------------------------------------------------
            # LIGNE COMMANDE (580010DS5)
            # Libellé = numéro de commande
            # ------------------------------------------------------------
            lignes.append({
                "compte": COMPTE_COMMANDE,
                "objet": str(commande) if commande is not None else "",
                "d": montant_fmt if debit_commande else "",
                "c": "" if debit_commande else montant_fmt,
                "date": date_ecriture,
                "piece": piece
            })

            # ------------------------------------------------------------
            # LIGNE AVOIR (580012DS5)
            # Libellé = Nom + Prénom + date d’expiration
            # ------------------------------------------------------------
            exp_str = (
                date_expiration.strftime("%d/%m/%Y")
                if date_expiration is not None else ""
            )
            libelle_avoir = f"{nom} {prenom} {exp_str}".strip()

            lignes.append({
                "compte": COMPTE_AVOIR,
                "objet": libelle_avoir,
                "d": "" if debit_commande else montant_fmt,
                "c": montant_fmt if debit_commande else "",
                "date": date_ecriture,
                "piece": piece
            })

        # Aucun mouvement détecté
        if not lignes:
            raise NotAvoirFileError("Aucune ligne AVOIR détectée")

        # ------------------------------------------------------------
        # EXPORT DU FICHIER COMPTABLE
        # ------------------------------------------------------------
        sortie = DOSSIER_SORTIE / f"{fichier.stem}_avoirs.csv"

        with open(sortie, "w", newline="", encoding="latin1") as f:
            writer = csv.writer(f, delimiter=";")

            # En-tête comptable
            writer.writerow([
                "STE", "DATE", "COMPTE", "Auxiliaire",
                "n° pièce", "OBJET", "D", "C",
                "Journal", "Analytique"
            ])

            # Lignes comptables
            for l in lignes:
                writer.writerow([
                    STE,
                    l["date"],
                    l["compte"],
                    AUXILIAIRE,
                    l["piece"],
                    l["objet"],
                    l["d"],
                    l["c"],
                    JOURNAL,
                    ANALYTIQUE
                ])

        return sortie

    finally:
        # Fermeture explicite du classeur Excel
        wb.close()
