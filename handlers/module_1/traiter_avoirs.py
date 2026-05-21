import csv
from pathlib import Path
from typing import Optional
import pandas as pd                                          # ← AJOUT
from openpyxl import load_workbook
from config import DOSSIER_SORTIE
from logger import logger
from core.moniteur_schema import comparer_schema             # ← AJOUT

# ==========================================================
# EXCEPTION MÉTIER
# ==========================================================

class NotAvoirFileError(Exception):
    """Levée si aucun avoir exploitable n'est détecté."""
    pass

# ==========================================================
# INDEX DES COLONNES (base 0)
# ==========================================================
COL_NOM            = 1   # B
COL_PRENOM         = 2   # C
COL_DATE_CREATION  = 4   # E
COL_DATE_EXPIR     = 5   # F
COL_STATUT_G       = 6   # G
COL_STATUT_H       = 7   # H  ← contient "Remboursement" ou "Avoir"
COL_MONTANT        = 8   # I
COL_COMMANDE       = 10  # K

# ==========================================================
# CONSTANTES COMPTABLES
# ==========================================================
STE             = "DLM"
COMPTE_COMMANDE = "580010DS5"
COMPTE_AVOIR    = "580012DS5"
JOURNAL         = "OD"
AUXILIAIRE      = ""
ANALYTIQUE      = ""

# ==========================================================
# UTILITAIRES
# ==========================================================

def format_montant(valeur) -> str:
    """
    Formate un montant pour l'export comptable.
    - Valeur absolue, 2 décimales, virgule décimale.
    """
    if valeur is None:
        return ""
    if isinstance(valeur, str):
        valeur = float(valeur.replace(",", ".").strip())
    return f"{abs(valeur):.2f}".replace(".", ",")


def _norm(val) -> str:
    """Normalise une valeur texte : strip + majuscules."""
    return str(val).strip().upper() if val is not None else ""


def _statut_to_debit_commande(statut_g, statut_h, montant) -> Optional[bool]:
    s1 = _norm(statut_g)
    s2 = _norm(statut_h)
    positif = montant >= 0

    if s2 == "AVOIR":
        # positif → Crédit 580010 / Débit 580012 → debit_commande = False
        # négatif → Crédit 580010 / Débit 580012 → debit_commande = False
        return False

    if s2 == "REMBOURSEMENT":
        # positif → Débit 580010 / Crédit 580012 → debit_commande = True
        # négatif → Crédit 580010 / Débit 580012 → debit_commande = False
        return positif

    # Fallback colonne G
    if s1 in ("AVOIR", "CONSUMED"):
        return False

    if s1 in ("DEDUCTED", "REMBOURSEMENT"):
        return positif

    logger.warning(f"Statut avoir inconnu : G={s1!r} H={s2!r} → ligne ignorée")
    return None

# ==========================================================
# HANDLER PRINCIPAL
# ==========================================================

def traiter_avoirs(fichier: Path) -> Optional[Path]:
    """
    Traite un fichier d'avoirs (.xlsx) et génère les écritures comptables.

    Structure de sortie :
    STE | DATE | COMPTE | Auxiliaire | n°pièce | OBJET | D | C | Journal | Analytique
    """

    fichier = Path(fichier)
    if not fichier.exists():
        raise FileNotFoundError(f"Fichier AVOIR introuvable : {fichier}")

    logger.info(f"Traitement AVOIRS : {fichier.name}")

    lignes     = []
    nb_ignores = 0

    wb = load_workbook(fichier, data_only=True)

    try:
        ws = wb.active

        # ----------------------------------------------------------
        # Vérification du schéma (snapshot ligne 1 via pandas)    # ← AJOUT
        # ----------------------------------------------------------
        df_schema = pd.read_excel(fichier, nrows=0, header=0)     # ← AJOUT
        comparer_schema(df_schema, "avoirs")                       # ← AJOUT

        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):

            # Lecture des cellules
            nom            = row[COL_NOM].value
            prenom         = row[COL_PRENOM].value
            date_creation  = row[COL_DATE_CREATION].value
            date_expir     = row[COL_DATE_EXPIR].value
            statut_g       = row[COL_STATUT_G].value
            statut_h       = row[COL_STATUT_H].value if len(row) > COL_STATUT_H else None
            montant        = row[COL_MONTANT].value
            commande       = row[COL_COMMANDE].value

            # Ligne incomplète
            if montant is None or date_creation is None:
                logger.debug(f"Ligne {idx} ignorée : montant ou date manquant")
                nb_ignores += 1
                continue

            # Sens comptable
            debit_commande = _statut_to_debit_commande(statut_g, statut_h, montant)
            if debit_commande is None:
                nb_ignores += 1
                continue

            # Formatage
            try:
                date_str = date_creation.strftime("%d/%m/%Y")
            except AttributeError:
                logger.warning(f"Ligne {idx} ignorée : date_creation non formatable ({date_creation!r})")
                nb_ignores += 1
                continue

            piece       = f"JOURNEE DU {date_str}"
            montant_fmt = format_montant(montant)

            exp_str = (
                date_expir.strftime("%d/%m/%Y")
                if date_expir is not None else ""
            )
            libelle_avoir = f"{_norm(nom)} {_norm(prenom)} {exp_str}".strip()
            libelle_cmd   = str(commande).strip() if commande is not None else ""

            # ✅ Ligne COMMANDE (580010DS5)
            lignes.append({
                "date": date_str,
                "piece": piece,
                "compte": COMPTE_COMMANDE,
                "objet": libelle_cmd,
                "d": montant_fmt if debit_commande else "",  # ← CORRIGÉ
                "c": "" if debit_commande else montant_fmt,  # ← CORRIGÉ
            })

            # ✅ Ligne AVOIR (580012DS5)
            lignes.append({
                "date": date_str,
                "piece": piece,
                "compte": COMPTE_AVOIR,
                "objet": libelle_avoir,
                "d": "" if debit_commande else montant_fmt,  # ← CORRIGÉ
                "c": montant_fmt if debit_commande else "",  # ← CORRIGÉ
            })
    
        # ----------------------------------------------------------
        # Aucun mouvement détecté
        # ----------------------------------------------------------
        if not lignes:
            raise NotAvoirFileError(
                f"Aucune ligne AVOIR exploitable dans {fichier.name}"
            )

        logger.info(
            f"{len(lignes) // 2} avoirs traités "
            f"({nb_ignores} lignes ignorées)"
        )

        # ----------------------------------------------------------
        # Export CSV
        # ----------------------------------------------------------
        DOSSIER_SORTIE.mkdir(parents=True, exist_ok=True)

        sortie = DOSSIER_SORTIE / f"{fichier.stem}_avoirs.csv"

        with open(sortie, "w", newline="", encoding="latin1") as f:
            writer = csv.writer(f, delimiter=";")

            writer.writerow([
                "STE", "DATE", "COMPTE", "Auxiliaire",
                "n°pièce", "OBJET", "D", "C",
                "Journal", "Analytique"
            ])

            for l in lignes:
                writer.writerow([
                    STE,
                    l["date"],
                    l["compte"],
                    AUXILIAIRE,
                    l["piece"],
                    l["objet"],
                    l["d"],
                    l["c"],
                    JOURNAL,
                    ANALYTIQUE,
                ])

        logger.info(f"Export AVOIRS : {sortie.name} ({len(lignes)} écritures)")
        return sortie

    finally:
        wb.close()
